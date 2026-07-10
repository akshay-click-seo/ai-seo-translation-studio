"""
services/excel_service.py — Excel translation service.

Capabilities:
  - List all sheets and columns in an uploaded workbook.
  - Translate one or more columns (user-selected).
  - Write translated values to a new column or replace in-place.
  - Preserve all cell formatting, formulas, merged cells.
  - Output a translated .xlsx file ready for download.
"""

import os
import io
from typing import Optional, Callable

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from logger import setup_logger, translation_logger
from utils import is_translatable, output_filename, save_output
import time

logger = setup_logger(__name__)


class ExcelService:
    """Handles Excel file translation with column-level control."""

    # ── Introspection ──────────────────────────────────────────────────────────

    def get_sheets(self, file_path: str) -> list[str]:
        """
        Return all sheet names in the workbook.

        Args:
            file_path: Path to the .xlsx file.

        Returns:
            List of sheet name strings.
        """
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        except Exception as exc:
            logger.error("Failed to read sheets: %s", exc)
            return []

    def get_columns(self, file_path: str, sheet_name: Optional[str] = None) -> list[str]:
        """
        Return column headers from the first row of the specified sheet.

        Args:
            file_path:   Path to the .xlsx file.
            sheet_name:  Sheet to inspect (None = first sheet).

        Returns:
            List of header strings.
        """
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name or 0, nrows=0)
            return list(df.columns.astype(str))
        except Exception as exc:
            logger.error("Failed to read columns: %s", exc)
            return []

    def preview(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        rows: int = 5,
    ) -> pd.DataFrame:
        """
        Return a preview DataFrame (first *rows* data rows).

        Args:
            file_path:   Path to the .xlsx file.
            sheet_name:  Sheet to preview.
            rows:        Number of data rows to include.

        Returns:
            Pandas DataFrame.
        """
        try:
            return pd.read_excel(file_path, sheet_name=sheet_name or 0, nrows=rows)
        except Exception as exc:
            logger.error("Preview failed: %s", exc)
            return pd.DataFrame()

    # ── Translation ────────────────────────────────────────────────────────────

    def translate_columns(
        self,
        file_path: str,
        columns: list[str],
        translator,
        target_lang: str = "es-la",
        mode: str = "Technical",
        sheet_name: Optional[str] = None,
        add_suffix: bool = True,
        use_tm: bool = True,
        on_progress: Optional[Callable[[int, int], None]] = None,
    ) -> bytes:
        """
        Translate selected columns in an Excel workbook.

        For each column in *columns*:
          - Reads all cell values in that column.
          - Translates each non-empty, translatable value.
          - Writes the translation to a new column named "<original>_ES"
            (or replaces in-place if add_suffix=False).

        Args:
            file_path:   Path to source .xlsx file.
            columns:     List of column header names to translate.
            translator:  Translator instance.
            target_lang: Target language code.
            mode:        Translation mode.
            sheet_name:  Target sheet (None = first sheet).
            add_suffix:  If True, write to new columns; otherwise overwrite.
            use_tm:      Use Translation Memory.
            on_progress: Progress callback(current_row, total_rows).

        Returns:
            Translated workbook as raw bytes (suitable for st.download_button).
        """
        t0 = time.perf_counter()

        # Load workbook preserving formatting
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active

        # Build header map: column_name → column_index (1-based)
        headers: dict[str, int] = {}
        max_col = ws.max_column or 1
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col_idx)
            if cell.value is not None:
                headers[str(cell.value)] = col_idx

        # Determine next available column for translated output
        next_col = max_col + 1
        translated_col_map: dict[str, int] = {}

        for col_name in columns:
            if col_name not in headers:
                logger.warning("Column '%s' not found in sheet.", col_name)
                continue

            if add_suffix:
                out_col = next_col
                out_header = f"{col_name}_ES"
                ws.cell(row=1, column=out_col, value=out_header)
                translated_col_map[col_name] = out_col
                next_col += 1
            else:
                translated_col_map[col_name] = headers[col_name]

        # Count total rows to translate
        total_rows = ws.max_row - 1  # exclude header
        translated_count = 0

        for row_idx in range(2, ws.max_row + 1):
            for col_name, src_col in headers.items():
                if col_name not in translated_col_map:
                    continue

                src_cell = ws.cell(row=row_idx, column=src_col)
                raw_value = src_cell.value

                if raw_value is None:
                    continue

                str_value = str(raw_value).strip()
                if not str_value or not is_translatable(str_value):
                    # Copy as-is
                    out_col = translated_col_map[col_name]
                    ws.cell(row=row_idx, column=out_col, value=raw_value)
                    continue

                try:
                    translation = translator.translate_cell(
                        str_value, target_lang=target_lang, use_tm=use_tm
                    )
                    out_col = translated_col_map[col_name]
                    ws.cell(row=row_idx, column=out_col, value=translation)
                    translated_count += 1
                except Exception as exc:
                    logger.error("Row %d col '%s' error: %s", row_idx, col_name, exc)
                    out_col = translated_col_map[col_name]
                    ws.cell(row=row_idx, column=out_col, value=raw_value)

            if on_progress:
                on_progress(row_idx - 1, total_rows)

        # Serialize to bytes
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        result_bytes = buf.read()

        elapsed = (time.perf_counter() - t0) * 1000
        translation_logger.log_file_translation(
            file_name=os.path.basename(file_path),
            file_type="xlsx",
            mode=mode,
            target_lang=target_lang,
            segments=translated_count,
            processing_time_ms=elapsed,
        )
        logger.info(
            "Excel translated: %d cells in %.0f ms", translated_count, elapsed
        )
        return result_bytes

    def translate_full_sheet(
        self,
        file_path: str,
        translator,
        target_lang: str = "es-la",
        mode: str = "Technical",
        sheet_name: Optional[str] = None,
        use_tm: bool = True,
        on_progress: Optional[Callable[[int, int], None]] = None,
    ) -> bytes:
        """
        Translate ALL text cells in the selected sheet.

        Args:
            file_path:   Path to the .xlsx file.
            translator:  Translator instance.
            target_lang: Target language code.
            mode:        Translation mode.
            sheet_name:  Sheet to translate (None = active sheet).
            use_tm:      Use Translation Memory.
            on_progress: Progress callback.

        Returns:
            Translated workbook bytes.
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active

        cells = [
            cell for row in ws.iter_rows() for cell in row
            if cell.value and isinstance(cell.value, str) and is_translatable(cell.value)
        ]
        total = len(cells)

        for idx, cell in enumerate(cells):
            if on_progress:
                on_progress(idx, total)
            try:
                translation = translator.translate_cell(
                    cell.value, target_lang=target_lang, use_tm=use_tm
                )
                cell.value = translation
            except Exception as exc:
                logger.error("Cell %s error: %s", cell.coordinate, exc)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()


# ── Module-level singleton ─────────────────────────────────────────────────────
excel_service = ExcelService()
